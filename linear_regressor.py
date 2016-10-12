import openpyxl
import numpy as np
import string
import random
import matplotlib.pyplot as p
from sklearn.svm import SVC
from sklearn import linear_model

wb = openpyxl.load_workbook('../spreadsheets/nice_sheet_aug_1.xlsx')
mainSheet = wb.active

MAX_ROW = 437
BXRESULT_H_COLUMN = "BY" # Matters only if you want to make sure to test only on people with BXRESULT = H
ANSWER_COLUMN = "P"
USE_LILI_DATA = False
TRAINING_SET_FRACTION = 0.5
VALIDATION_SET_FRACTION = 0.25
TEST_SET_FRACTION = 0.25
FALSE_POSITIVE_PENALTY = 1
FALSE_NEGATIVE_PENALTY = 10

assert TRAINING_SET_FRACTION + VALIDATION_SET_FRACTION + TEST_SET_FRACTION == 1

def extractDataList(sheet):
    dataListWithH = []
    dataListWithoutH = []
    
    for i in range(2, MAX_ROW+1):
        newDataPoint = [[]]
        j = 0
        currentColumnName = "A"
    
        currentHNess = "error! error! error! AAAAAAAAUUUGH!"
    
        while not currentColumnName == ANSWER_COLUMN:
            if currentColumnName == BXRESULT_H_COLUMN:
                currentHNess = (mainSheet[currentColumnName + str(i)].value == 1)
        
            newDataPoint[0].append(float(mainSheet[currentColumnName + str(i)].value))
        
            j += 1
            currentColumnName = convertToColumnLabel(j)
        
        newDataPoint.append([float(mainSheet[ANSWER_COLUMN + str(i)].value)])
        
        if currentHNess:    
            dataListWithH.append(newDataPoint)    
        else:
            dataListWithoutH.append(newDataPoint)        
            
    return dataListWithH, dataListWithoutH

def f1Score(truePositives, trueNegatives, falsePositives, falseNegatives):
#    print truePositives, trueNegatives, falsePositives, falseNegatives

    if truePositives == 0:
        return 0
    
    precision = float(truePositives) / (falsePositives + truePositives)
    recall = float(truePositives) / (truePositives + falseNegatives)
    
    return 2*(precision*recall)/(precision+recall)

def convertToColumnLabel(x):
    lastDigit = string.ascii_uppercase[x % 26]
        
    dividend = x - (x%26)
    
    if dividend == 0:
        return lastDigit
    
    else:
        return convertToColumnLabel(dividend/26 - 1) + lastDigit        

def turnDataListIntoTwoArrays(dataList):
    x = np.array([i[0] for i in dataList])
    y = np.array([i[1] for i in dataList])
    
    return x,y

def turnDataListIntoArrayAndList(dataList):
    x = np.array([i[0] for i in dataList])
    y = [i[1][0] for i in dataList]
    
    return x,y

def printCoeffsForWholeDataSet(dataList):
#    eiWB = openpyxl.load_workbook('../spreadsheets/data_simple_processed_extra_info.xlsx')
#    eiSheet = eiWB.get_sheet_by_name(name = 'extra_info') 
    
    x, y = turnDataListIntoTwoArrays(dataList)

    regr = linear_model.LinearRegression()
    regr.fit(x, y)

    listOfCoeffs = []

    for j, coef in enumerate(regr.coef_[0]):
        listOfCoeffs.append((mainSheet[convertToColumnLabel(j) + "1"].value, coef))
        listOfCoeffs.sort(key=lambda x: abs(x[1]))
        listOfCoeffs.reverse()
    
    output = open("coeffs.txt", "w")
    
    for t in listOfCoeffs:
        output.write(t[0] + " " + str(t[1]) + "\n")
        
def printCoeffsForWholeDataSetExtraInfo(dataList):
    eiWB = openpyxl.load_workbook('../spreadsheets/data_simple_processed_extra_info.xlsx')
    eiSheet = eiWB.get_sheet_by_name(name = 'extra_info') 
    
    x, y = turnDataListIntoTwoArrays(dataList)

    regr = linear_model.LinearRegression()
    regr.fit(x, y)
    
    ENOUGH_TO_INCLUDE = 20
    
    includeDict = {}
    
    j = 0
    currentColumnName = "A"
    
    while not currentColumnName == ANSWER_COLUMN:
        columnLabel = eiSheet[currentColumnName + "1"].value
        columnSum = eiSheet[currentColumnName + str(MAX_ROW + 1)].value
        
        includeDict[columnLabel] = (columnSum > ENOUGH_TO_INCLUDE) and (not currentColumnName == BXRESULT_H_COLUMN)
        
        j += 1
        currentColumnName = convertToColumnLabel(j)

    listOfCoeffs = []

    for j, coef in enumerate(regr.coef_[0]):
        listOfCoeffs.append((mainSheet[convertToColumnLabel(j) + "1"].value, coef))
        listOfCoeffs.sort(key=lambda x: abs(x[1]))
        listOfCoeffs.reverse()
    
    output = open("coeffs.txt", "w")
    
    print includeDict
    
    for t in listOfCoeffs:
        if t[0] in includeDict and includeDict[t[0]]:
            output.write(t[0] + " " + str(t[1]) + "\n")

def train(regr, trainingSet):
    x, y = turnDataListIntoTwoArrays(trainingSet)
    regr.fit(x, y)

def findOptimalThresholds(regr, validationSet):
    x, y = turnDataListIntoArrayAndList(validationSet)

    predictions = [i[0] for i in regr.predict(x)]
    
    truePositiveList = []
    trueNegativeList = []
    falsePositiveList = []
    falseNegativeList = []
    thresholdList = []
    
    for thresholdTimesAHundred in range(-100, 200):
    
        threshold = thresholdTimesAHundred/100.

        ourAnswers = [i > threshold for i in predictions]
    
        truePositives = sum([i and j for i, j in zip(ourAnswers, y)])
        trueNegatives = sum([(not i) and (not j) for i, j in zip(ourAnswers, y)])
        falsePositives = sum([i and (not j) for i, j in zip(ourAnswers, y)])
        falseNegatives = sum([(not i) and j for i, j in zip(ourAnswers, y)])
    
        truePositiveList.append(truePositives)
        trueNegativeList.append(trueNegatives)
        falsePositiveList.append(falsePositives)
        falseNegativeList.append(falseNegatives)
        thresholdList.append(threshold)
    
    f1Scores = [f1Score(tp, tn, fp, fn) for tp, tn, fp, fn in zip(truePositiveList, \
        trueNegativeList, falsePositiveList, falseNegativeList)]    
        
    bestF1threshold = thresholdList[f1Scores.index(max(f1Scores))]

    falseNegativeList.reverse()
    bestNoCancerThresholdReversed = falseNegativeList.index(0)
    falseNegativeList.reverse()
    bestNoCancerThreshold = thresholdList[len(thresholdList) - bestNoCancerThresholdReversed - 1]
    
    print f1Scores
    
    print bestF1threshold, bestNoCancerThreshold
    return bestF1threshold, bestNoCancerThreshold

def test(regr, testSet, threshold):
    x, y = turnDataListIntoArrayAndList(testSet)
    predictions = [i[0] for i in regr.predict(x)]
    
    ourAnswers = [i > threshold for i in predictions]
    
    truePositives = sum([i and j for i, j in zip(ourAnswers, y)])
    trueNegatives = sum([(not i) and (not j) for i, j in zip(ourAnswers, y)])
    falsePositives = sum([i and (not j) for i, j in zip(ourAnswers, y)])
    falseNegatives = sum([(not i) and j for i, j in zip(ourAnswers, y)])

    return truePositives, trueNegatives, falsePositives, falseNegatives
        
if USE_LILI_DATA:
    liliWb = openpyxl.load_workbook("../spreadsheets/lili_train_processed.xlsx")
    liliSheet = liliWb.active
    
    dataListWithH, dataListWithoutH = extractDataList(liliSheet)
    
    trainingSetSize = int(TRAINING_SET_FRACTION*len(dataListWithH)) 
    validationSetSize = int(VALIDATION_SET_FRACTION*len(dataListWithH))

    trainingSetWithH = dataListWithH[:trainingSetSize]
    validationSet = dataListWithH[trainingSetSize:trainingSetSize+validationSetSize]
    testSet = dataListWithH[trainingSetSize+validationSetSize:]
    
else:
    
    dataListWithH, dataListWithoutH = extractDataList(mainSheet)
    
    random.shuffle(dataListWithH)
    random.shuffle(dataListWithoutH)

    trainingSetSize = int(TRAINING_SET_FRACTION*len(dataListWithH)) 
    validationSetSize = int(VALIDATION_SET_FRACTION*len(dataListWithH))

    trainingSetWithH = dataListWithH[:trainingSetSize]
    validationSet = dataListWithH[trainingSetSize:trainingSetSize+validationSetSize]
    testSet = dataListWithH[trainingSetSize+validationSetSize:]

trainingSet = trainingSetWithH + dataListWithoutH

regr = linear_model.LinearRegression()

train(regr, trainingSet)
bestF1threshold, bestNoCancerThreshold = findOptimalThresholds(regr, validationSet)

tpF1, tnF1, fpF1, fnF1 = test(regr, testSet, bestF1threshold)
tpNC, tnNC, fpNC, fnNC = test(regr, testSet, bestNoCancerThreshold)

print "Using optimal F1 threshold of", bestF1threshold, "over validation set:"
print "F1 SCORE:", f1Score(tpF1, tnF1, fpF1, fnF1)
print "True positives:", tpF1
print "True negatives:", tnF1
print "False positives:", fpF1
print "False negatives:", fnF1
print ""
print "Using optimal no-cancer threshold of", bestNoCancerThreshold, "over validation set:"
print "F1 SCORE:", f1Score(tpNC, tnNC, fpNC, fnNC)
print "True positives:", tpNC
print "True negatives:", tnNC
print "False positives:", fpNC
print "False negatives:", fnNC
    
#numBenigns = falsePositiveList[0]
#numMalignancies = falseNegativeList[-1]
#
#for i in range(MAX_MALIGNANCY_MISSES+1):
#    lastNumber = "error! error! error!"
#    
#    for j in range(len(falsePositiveList)):
#        
#        if falseNegativeList[j] > i:
#            print "If we are willing to tolerate", str(i) + "/" + str(numMalignancies), \
#                "missed malignancies, we can save", str(numBenigns - lastNumber) + "/" + str(numBenigns), \
#                "patients who don't have cancer from getting surgery."
#                
#            break
#                
#        lastNumber = falsePositiveList[j]    
#    
#p.plot(thresholdList, falsePositiveList, "b-")
#p.plot(thresholdList, falseNegativeList, "r-")
#
#p.savefig("false_pos_neg.png")
#
#p.clf()
#
#penalties = [FALSE_POSITIVE_PENALTY*i + FALSE_NEGATIVE_PENALTY*j \
#    for i,j in zip(falsePositiveList, falseNegativeList)]
#    
#p.plot(thresholdList, penalties, "b-")
#
#p.savefig("penalties.png")

printCoeffsForWholeDataSetExtraInfo(trainingSet)

#print vectorizer

#print wb["all"]["A1"].value