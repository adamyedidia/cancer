
# coding: utf-8

# In[1]:

from __future__ import division
import json, glob, os, numpy, sys
import matplotlib
from data_process_functions_lili_original import *
os.chdir("/Users/adamyedidia/breast_cancer/spreadsheets/")

import openpyxl, pprint

matplotlib.rcParams.update({'font.size': 10})

SHEET_NAME = 'manisha_new.xlsx'
#LIST_OF_BINARY_COLS = ['Q', 'R', 'S', 'T','U','V','W','X','Y','Z','AA','AB','AC']
LIST_OF_BINARY_COLS = ['D','F','G','H','I','J','K','L','M','T','U','W','X','Y','AA','AE','AH','AJ']
LIST_OF_REAL_COLS = ['A','C','N','O','P','Q','R','S','V','Z','U','V','Z']
ANSWER_COL = 'AQ'

data,feature_names,result=dataprocess(SHEET_NAME,LIST_OF_BINARY_COLS,LIST_OF_REAL_COLS,ANSWER_COL)
yy=[1 if x=='M' else 0 for x in result]
print 'cancer numbers: %d' %yy.count(1)
y=np.array(yy)
X_structure=data[:,1:]
feature_names_structure=feature_names[1:]
n=X_structure.shape[0]


def readreport(SHEET_NAME):
    sheet = openpyxl.load_workbook(SHEET_NAME).active
    report=[]
    for j in range (1,sheet.max_column+1):
        if sheet.cell(row=1,column=j).value=='BXPATHREP':
            for i in range(2,sheet.max_row+1):
                rpt=sheet.cell(row=i,column=j).value
                #print rpt
                #rpt.replace('_x000D_\n_x000D_\n', '\n')
                rpt=rpt.replace('_x000D_\n_x000D_\n', '')
                rpt=rpt.replace('MASSACHUSETTS GENERAL HOSPITAL','')
                rpt=rpt.replace('BOSTON, MA  REMOVED_CASE_ID','')
                rpt=rpt.lower()
                rpt=rpt.replace('removed_patient_name','')
                rpt=rpt.replace('_patient_name_removed','')
                rpt=rpt.replace('removed_accession_id','')
                rpt=rpt.replace('removed_case_id','')
                rpt=rpt.replace('removed_date','')
                rpt=rpt.replace('removed_month','')
                rpt=rpt.replace('removed_patient_','')
                rpt=rpt.replace('\n\n','')
                rpt=rpt.replace('\n \n','')
                rpt.strip('\n\n')
                rpt.strip('\n \n')
                report.append(rpt)
    return report

report= readreport(SHEET_NAME)
vectorizer = CountVectorizer(ngram_range=(1, 2),min_df=1)
#XXX = vectorizer.fit_transform(report)
ft_report= numpy.array(vectorizer.fit_transform(report).toarray())

feature_names_report = vectorizer.get_feature_names()
feature_names_report=np.array(feature_names_report,dtype=str)
print "total uni & bi gram features are :%s" %str(feature_names_report.shape)

count=numpy.sum(ft_report,axis=0)
delete=[]
thresh=20             #####################===============you can change this guy
for i in range(1,count.shape[0]):
    if count[i]<thresh:
        delete.append(i)
ft_report_p=numpy.delete(ft_report,delete,1)
feature_names_report_prune=numpy.delete(feature_names_report,delete,0)
print "total uni & bi gram features pruned with threshold of 20  are  :%s" %str(feature_names_report_prune.shape)
print feature_names_report_prune[:10]
ft_report_prune=ft_report_p


# In[2]:

ft_report_prune.shape


# In[3]:

ft_report_prune.shape


# In[4]:

f_structure= ExtraTreesClassifier(n_estimators=200,max_depth=12,max_features=80,class_weight = {1:7},
                              random_state=123)
f_report = ExtraTreesClassifier(n_estimators=200,max_depth=10,max_features=500,class_weight = {1:5},
                              random_state=123)
#feature_ranking(f_structure,feature_names_structure,X_structure,y,'importantfeature_structure.txt',pp=1)
#feature_ranking(f_report,feature_names_report_prune,ft_report_prune,y,'importantfeature_report.txt',pp=1)


# In[5]:

from sklearn.cross_validation import StratifiedShuffleSplit

n_str=100
n_report=180
weight =3
random_state= 127

#sss = StratifiedShuffleSplit(y, 10, test_size=0.05, random_state=1237)
#rs = cross_validation.ShuffleSplit(len(y), n_iter=10,test_size=.1, random_state=1237)

from sklearn.cross_validation import StratifiedKFold
skf = StratifiedKFold(y, 6, shuffle=True, random_state=random_state)

k=0

YTEST=[]
YTEST_index=[]
Probability=[]

def describeResults(mResults, bResults, numMResults, numBResults):
    bestLowRisk = 0
    bestMedRisk = 0
    bestHighRisk = 0
    for i, mResult in enumerate(mResults):
        if mResult/float(numMResults) > 0.95:
            bestLowRisk = max(bestLowRisk, bResults[i])
        elif mResult/float(numMResults) > 0.80:
            bestMedRisk = max(bestMedRisk, bResults[i])
        else:
            bestHighRisk = max(bestHighRisk, bResults[i])
    bestMedRisk -= bestLowRisk
    bestHighRisk = numBResults - bestMedRisk - bestLowRisk
    bestLowRiskFrac = bestLowRisk / float(numBResults)
    bestMedRiskFrac = bestMedRisk / float(numBResults)
    bestHighRiskFrac = bestHighRisk / float(numBResults)
    print 'BEST LOW RISK PRECISION:'
    print str(bestLowRiskFrac*100) + '%', bestLowRisk, 'patients'
    print 'BEST MEDIUM RISK PRECISION:'
    print str(bestMedRiskFrac*100) + '%', bestMedRisk, 'patients'
    print 'BEST HIGH RISK PRECISION:'
    print str(bestHighRiskFrac*100) + '%', bestHighRisk, 'patients'

overallMResults = [0]*21
overallBResults = [0]*21
numBResultsTotal = 0
numMResultsTotal = 0

tSetCounter = 0

sumGroupsToCounts = {"low": [0,0],
                  "med": [0,0],
                  "high": [0,0]}

# [TP, FP, TN, FN]
categoryThresholds = {0.05: [0,0,0,0],
                      0.15: [0,0,0,0],
                      0.3: [0,0,0,0]}

for train_index, test_index in skf:
        ff = ExtraTreesClassifier(n_estimators=200,max_depth=10,max_features=100
                                      ,class_weight = {1:weight},random_state=123)

        print 'new sets'
        X_train_struct, X_test_struct = X_structure[train_index], X_structure[test_index]
        X_train_txt, X_test_txt = ft_report_prune[train_index], ft_report_prune[test_index]
        print X_train_txt[:10]
        y_train, y_test = y[train_index], y[test_index]

        YTEST.append(y_test)
        YTEST_index.append(test_index)

        k+=1
#        feature_ranking(f_structure,feature_names_structure,X_train_struct,y_train,'importantfeature_strcture_'+str(k)+'.txt',pp=0)
#        feature_ranking(f_report,feature_names_report_prune,X_train_txt,y_train,'importantfeature_report_'+str(k)+'.txt',pp=0)

        indx_structure=get_sorted_ft_idx(f_structure,X_train_struct,y_train,n=n_str)
        indx_txt=get_sorted_ft_idx(f_report,X_train_txt,y_train,n=n_report)

        X_train=np.concatenate((X_train_struct[:,indx_structure], X_train_txt[:,indx_txt]), axis=1)
        X_test=np.concatenate((X_test_struct[:,indx_structure], X_test_txt[:,indx_txt]), axis=1)


        ff.fit(X_train, y_train)

        feature_names = vectorizer.get_feature_names()
        importances = ff.feature_importances_
        print 'lens', len(importances), len(feature_names)
        std = np.std([tree.feature_importances_ for tree in ff.estimators_], axis=0)
        indices = np.argsort(importances)[::-1]
        #print('Feature ranking:')
        #for f in range(X_train.shape[1]):
        #    print('%d. feature %d (%f)' % (f + 1, indices[f], importances[indices[f]]))
        #for f in range(X_train.shape[1]):
        #    print('%d. feature %s (%f)' % (f + 1, feature_names[indices[f]], importances[indices[f]]))
        ff_proba=ff.predict_proba(X_test)
        Probability.append(ff_proba)
        indx1=np.where(y_test==1)
        indx2=np.where(y_test==0)
        #for i, val in enumerate(X_test_struct[indx1].tolist()):
        #    print i
        #    print val
        #print y_test
        print indx1
        #print ff_proba[:,1][indx1]
        def tryThreshold(thresh):
            mResult = 0
            bResult = 0
            for i in ff_proba[:,1][indx1]:
                if i>thresh:
                    mResult += 1
            for i in ff_proba[:,1][indx2]:
                if i<thresh:
                    bResult += 1
            return mResult, bResult



        def countPatientsWithCancer():
            namesToRanges = {"low": (0.0, 0.05),
                             "med": (0.05, 0.15),
                             "high": (0.15, 1.0)}


            groupsToCounts = {"low": [0,0],
                              "med": [0,0],
                              "high": [0,0]}


            # cancers
            for i in ff_proba[:,1][indx1]:
                for binName in ["low", "med", "high"]:
                    binRange = namesToRanges[binName]
                    if binRange[0] < i and i < binRange[1]:
                        groupsToCounts[binName][0] += 1

                # [TP, FP, TN, FN]
                for thresh in categoryThresholds:
                    if i < thresh:
                        categoryThresholds[thresh][3] += 1
                    else:
                        categoryThresholds[thresh][0] += 1

            # no-cancers
            for i in ff_proba[:,1][indx2]:
                for binName in ["low", "med", "high"]:
                    binRange = namesToRanges[binName]
                    if binRange[0] < i and i < binRange[1]:
                        groupsToCounts[binName][1] += 1

                # [TP, FP, TN, FN]
                for thresh in categoryThresholds:
                    if i < thresh:
                        categoryThresholds[thresh][2] += 1
                    else:
                        categoryThresholds[thresh][1] += 1

            return groupsToCounts

        tSetCounter += 1
        print "training set number", tSetCounter

        cpwc = countPatientsWithCancer()
        print cpwc
        for i in cpwc:
            print i, cpwc[i][0] / (cpwc[i][0] + cpwc[i][1]+0.000001), cpwc[i][0], "/", \
                cpwc[i][0] + cpwc[i][1]

        sumGroupsToCounts["low"][0] += cpwc["low"][0]
        sumGroupsToCounts["low"][1] += cpwc["low"][1]
        sumGroupsToCounts["med"][0] += cpwc["med"][0]
        sumGroupsToCounts["med"][1] += cpwc["med"][1]
        sumGroupsToCounts["high"][0] += cpwc["high"][0]
        sumGroupsToCounts["high"][1] += cpwc["low"][1]

        mResults = []
        bResults = []
        thresholds = []
        numMResults = float(len(ff_proba[:,1][indx1]))
        numBResults = float(len(ff_proba[:,1][indx2]))
        for threshLarge in range(21):
            thresh = threshLarge/100.
            mResult, bResult = tryThreshold(thresh)
 #           print mResult, bResult, numMResults, numBResults
            mResults.append(mResult)
            if mResult > 0.95:
                lowThresh = thresh
            if mResult > 0.80:
                medThresh = thresh

            bResults.append(bResult)
            thresholds.append(thresh)
            overallMResults[threshLarge] += mResult
            overallBResults[threshLarge] += bResult
        yy= np.random.random_sample((len(y_test),))
        plt.figure(1)
        subplotNum = 230 + tSetCounter
        print subplotNum
        plt.subplot(subplotNum)
        if tSetCounter == 1:
            plt.xlabel('Probability Threshold')
            plt.ylabel('Accuracy')
        plt.plot(thresholds, [mResult / numMResults for mResult in mResults], 'r-')
        plt.plot(thresholds, [bResult / numBResults for bResult in bResults], 'b-')
        numBResultsTotal += numBResults
        numMResultsTotal += numMResults
        ax = plt.gca()
        ax.set_xlim([0, 0.2])
        describeResults(mResults, bResults, numMResults, numBResults)
        #plt.scatter([0, 1], [0, 1])
        #plt.scatter(ff_proba[:,1][indx1], yy[indx1], c=y_test[indx1], s=50,marker='o', color='r')
        #plt.scatter(ff_proba[:,1][indx2], yy[indx2], c=y_test[indx2], s=50,marker='+', color='b')
        plt.figure(2)
        plt.subplot(subplotNum)
        for i in indx1:
            plt.plot(ff_proba[:,1][i], yy[i], 'ro')
        for i in indx2:
            plt.plot(ff_proba[:,1][i], yy[i], 'b+')
        ax = plt.gca()
        ax.set_xlim([0, 1.])
        if tSetCounter == 1:
            plt.xlabel('Probability')
            plt.ylabel('Random number')
        plt.legend(loc='best')

        plt.figure(3)
        for i in indx1:
            plt.plot(ff_proba[:,1][i], yy[i], 'ro')
        for i in indx2:
            plt.plot(ff_proba[:,1][i], yy[i], 'b+')
#        plt.show()
        #prb=cus_probability(threshold, ff_proba[:,1])
        #prb=np.array(prb)
        #break

for i in sumGroupsToCounts:
    print i, sumGroupsToCounts[i][0] / (sumGroupsToCounts[i][0] + sumGroupsToCounts[i][1]), \
        sumGroupsToCounts[i][0], "/", sumGroupsToCounts[i][0] + sumGroupsToCounts[i][1]

plt.figure(1)
plt.savefig("six_graphs.png", format="png", dpi=1200)

plt.figure(2)
plt.savefig("six_scatters.png", format="png", dpi=1200)

plt.figure(3)

plt.xlabel('Probability')
plt.ylabel('Random number')
plt.legend(loc='best')
ax = plt.gca()
ax.set_xlim([0, 1.])

plt.axvspan(0, 0.05, alpha=0.5, color='#CDFAE7') # green
plt.axvspan(0.05, 0.15, alpha=0.5, color='#F6FACD') # yellow
plt.axvspan(0.15, 1., alpha=0.5, color='#FACDE0') # red

plt.savefig("aggregate_scatters.png", format="png", dpi=1200)

plt.figure(4)
plt.xlabel('Probability Threshold')
plt.ylabel('Accuracy')
for mResult in overallMResults:
    if mResult/numMResultsTotal > 0.95:
        lowThresh = thresh
    if mResult/numMResultsTotal > 0.80:
        medThresh = thresh
print lowThresh, medThresh
plt.plot(thresholds, [mResult / numMResultsTotal for mResult in overallMResults], 'r-')
plt.plot(thresholds, [bResult / numBResultsTotal for bResult in overallBResults], 'b-')
plt.axvspan(0, 0.05, alpha=0.5, color='#CDFAE7') # green
plt.axvspan(0.05, 0.15, alpha=0.5, color='#F6FACD') # yellow
plt.axvspan(0.15, 0.2, alpha=0.5, color='#FACDE0') # red

ax = plt.gca()
ax.set_xlim([0, 0.2])

plt.savefig('aggregate_graph.png', format='png', dpi=1200)

print categoryThresholds


sys.exit()
plt.show()


# In[6]:

p_all=[]
f_all=[]
r_all=[]
roc_all=[]
patients=[]
for ii in range(len(YTEST)):
    y_test=YTEST[ii]
    ff_proba=Probability[ii]
    roc=[]
    p=[]
    r=[]
    f=[]
    #results=[]
    pat=[]
    Thresh=[]
    for i in range(30):
        threshold=float(float(i+1)/80)
        Thresh.append(threshold)
        predicted=cus_predict(threshold, ff_proba[:,1])
        #results.append(predicted)

        fpr_rf, tpr_rf, _ = roc_curve(y_test,predicted)
        roc.append([fpr_rf, tpr_rf])
        p.append(precision_score(y_test,predicted))
        r.append(recall_score(y_test,predicted))
        f.append(f1_score(y_test,predicted))
        print(classification_report(y_test,predicted) )
        print 'roc_auc is %f :' %roc_auc_score(y_test,predicted)
        pt=len(predicted)-np.count_nonzero(predicted)
        print 'non cancer patient : %d' %pt
        pat.append(pt)

    #plt.scatter(prb[:,1],y_test)

    '''
    indxFb=np.intersect1d(np.where(ff_proba[:,1]>0.7 ),np.where(y_test==0))
    indxFm=np.intersect1d(np.where(ff_proba[:,1]<0.3),np.where(y_test==1))
    fp= X_test[indxFm][:,0]  ### The first column is the patient ID
    fn=X_test[indxFb][:,0]
    falsepositive.extend(fp)
    falsenegative.extend(fn)
    '''
    p_all.append(p)
    r_all.append(r)
    f_all.append(f)
    roc_all.append(roc)
    patients.append(pat)


# In[7]:

n=2
#m=f_all.shape[1]
m=10

f_all=np.array(f_all)
r_all=np.array(r_all)
patients=np.array(patients)
print 'trainging size: %d' %f_all[:-n,i].shape
for i in range(m):
    print average(f_all[:-n,i])
print "  \n test result"
for i in range(m):
    print average(f_all[-n:,i])
print " \n train result"
for i in range(m):
    print average(r_all[:-n,i])
print "  \ntest result"

for i in range(m):
    print average(r_all[-n:,i])

print "\n train result"
t=sum([len(l) for l in YTEST_index[:-n] ])
for i in range(m):
    print float(sum(patients[:-n,i])/float(sum([len(l) for l in YTEST_index[:-n] ])))

print ' \n test result'
for i in range(m):
    print float(sum(patients[-n:,i])/float(sum([len(l) for l in YTEST_index[-n:] ])))

patients[9,3]


# In[ ]:




# In[ ]:




# In[ ]:




# In[ ]:

#[54486, 86819,111803]
TestID=54486
newdata=[]
newtxt=[]
ID=[]
newy=[]
for i in range(data.shape[0]):
    if data[i,0] ==TestID:
        ID.append(data[i,0])
        newdata.append(data[i,1:])
        newtxt.append(ft_report_p[i,:])
        newy.append(y[i])
    '''
    else:
        if yy[i]==1:
            ID.append(data[i,0])
            newdata.append(data[i])
            newtxt.append(ft_report_p[i,:])
            newy.append(yy[i])
    '''
newdata=np.array(newdata)
newtxt=np.array(newtxt)
newy=np.array(newy)


# In[ ]:

X_test_special=np.concatenate((newdata[:,indx_structure], newtxt[:,indx_txt]), axis=1)
ff_proba=ff.predict_proba(X_test_special)
Probability.append(ff_proba)
indx1=np.where(newy==1)
indx2=np.where(newy==0)
#print indx1
yy= np.random.random_sample((len(newy),))
plt.clf()
#plt.scatter([0, 1], [0, 1])
plt.scatter(ff_proba[:,1][indx1], yy[indx1], c=newy[indx1], s=50,marker='o',label='cancer')
plt.scatter(ff_proba[:,1][indx2], yy[indx2], c=newy[indx2], s=50,marker='+',label='no cancer')
plt.xlabel('probability')
plt.ylabel('Random number')
plt.legend(loc='best')
plt.show()
#prb=cus_probability(threshold, ff_proba[:,1])
#prb=np.array(prb)


# In[ ]:




# In[ ]:




# In[ ]:




# In[ ]:




# In[ ]:




# In[52]:

print str(X_test.shape)
print len( indx_structure)
print len( indx_txt)


# In[ ]:
