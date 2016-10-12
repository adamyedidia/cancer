import openpyxl, pprint
import numpy as np
import string
import sys
import numbers

try:
    SHEET_NAME = sys.argv[1]
except:
    raise Exception("\n\nUsage: python data_processor.py [spreadsheet name].xlsx\n" + \
        "Puts the processed spreadsheet in [spreadsheet name]_processed.xlsx")

wbIn = openpyxl.load_workbook(SHEET_NAME)

MAX_ROW = 437

#LIST_OF_BINARY_COLS = ["C", "E", "F", "G", "H", "M", "N", "O", "P", "Q", "R", "S", 
#    "T", "V"]
    
LIST_OF_BINARY_COLS = ["Q", "R", "S", "T", "U", "W"]        
    
#LIST_OF_REAL_COLS = ["B", "I", "J", "K", "L"]

LIST_OF_REAL_COLS = []

ANSWER_COL = "AD"

FEATURE_EXCLUSION_THRESHOLD = 10



def convertToColumnLabel(x):
    lastDigit = string.ascii_uppercase[x % 26]
        
    dividend = x - (x%26)
    
    if dividend == 0:
        return lastDigit
    
    else:
        return convertToColumnLabel(dividend/26 - 1) + lastDigit        

def addToNumInstances(numInstancesDict, v):
    if v in numInstancesDict:
        numInstancesDict[v] += 1
    else:
        numInstancesDict[v] = 1

def addToValueList(differentValueDict, numInstancesDict, v, i):
    if v != None and v != "None" and string.replace(v, " ", "") != "" and v != "OTH" and v != "OTHE" and v != "OTHER":
        if v in differentValueDict:
            differentValueDict[v].append(i)  
            addToNumInstances(numInstancesDict, v)
        else:
            differentValueDict[v] = [i]
            addToNumInstances(numInstancesDict, v)

def getDifferentValueDict(wb, columnName):
    differentValueDict = {}
    numInstancesDict = {}
    
    mainSheet = wb.active
    
    columnDescriptor = mainSheet[columnName + "1"].value
    
    print columnDescriptor
    
    for i in range(2, MAX_ROW+1):
        cellName = columnName + str(i)
        cellValue = str(mainSheet[cellName].value)
        
        differentValues = string.split(cellValue, ",")
        
        for v in differentValues:
            addToValueList(differentValueDict, numInstancesDict, v, i)
            
    deletionList = []        
    for v in differentValueDict:
        if numInstancesDict[v] < FEATURE_EXCLUSION_THRESHOLD:
            print "Deleting feature", columnDescriptor + "_" + v, \
                "with only", str(numInstancesDict[v]), "instances."
            deletionList.append(v)
    
    for item in deletionList:
        del differentValueDict[item]
    
    return differentValueDict, columnDescriptor

def binarifyColumn(wb, columnName):
    differentValueDict = getDifferentValueDict(wb, columnName)
            
    newSheet = wb.create_sheet()
    newSheet.title = columnDescriptor    
    
    for i in range(2, MAX_ROW+1):        
        for j in range(len(differentValueDict)):
            newSheet[convertToColumnLabel(j) + str(i)].value = 0
            
    for j, v in enumerate(differentValueDict.keys()):        
        newSheet[convertToColumnLabel(j) + "1"].value = columnDescriptor + "_" + v
        
        for i in differentValueDict[v]:
            newSheet[convertToColumnLabel(j) + str(i)].value = 1    

def binarifyColumnSameSheet(wbIn, wbOut, columnName, columnCount):
    differentValueDict, columnDescriptor = getDifferentValueDict(wbIn, columnName)
    
    imaginarySheet = {}
    newSheet = wbOut.active
        
    for i in range(2, MAX_ROW+1):        
        for j in range(columnCount, columnCount+len(differentValueDict)):
            newSheet[convertToColumnLabel(j) + str(i)].value = 0
            
    for j, v in enumerate(differentValueDict.keys()):        
        newSheet[convertToColumnLabel(j+columnCount) + "1"].value = columnDescriptor + "_" + v
        
        for i in differentValueDict[v]:
            newSheet[convertToColumnLabel(j+columnCount) + str(i)].value = 1    
            
    return columnCount+len(differentValueDict)
    
def realifyColumnSameSheet(wbIn, wbOut, columnName, columnCount):
    valueSum = 0
    numValues = 0.
        
    inSheet = wbIn.active
    outSheet = wbOut.active
    
    columnDescriptor = inSheet[columnName + "1"].value    
        
    print columnDescriptor    
        
    for i in range(2, MAX_ROW+1):
        cellValue = inSheet[columnName + str(i)].value
        if isinstance(cellValue, numbers.Number):
            valueSum += cellValue
            numValues += 1.
            
    averageValue = valueSum/numValues
    
    outSheet[convertToColumnLabel(columnCount) + "1"].value = columnDescriptor
    
    for i in range(2, MAX_ROW+1):
        cellValue = inSheet[columnName + str(i)].value
        if isinstance(cellValue, numbers.Number):
            outSheet[convertToColumnLabel(columnCount) + str(i)].value = inSheet[columnName + str(i)].value
        else:
            if cellValue:
                print "Suspicious value", cellValue, "found on row", i, "of column", columnDescriptor + "."
                print "Value ignored."
            outSheet[convertToColumnLabel(columnCount) + str(i)].value = averageValue
            
    return columnCount + 1
        
def copyResultColumn(wbIn, wbOut, columnName, columnCount):
    inSheet = wbIn.active
    outSheet = wbOut.active
        
    for i in range(1, MAX_ROW+1):
        cellValue = inSheet[columnName + str(i)].value
        outSheet[convertToColumnLabel(columnCount) + str(i)].value = inSheet[columnName + str(i)].value
    
    print inSheet[columnName + "1"].value
    
    return columnCount + 1
    
wbOut = openpyxl.Workbook()
columnCount = 0

print ""
print "Binary Features:"
print ""

for binaryCol in LIST_OF_BINARY_COLS:
    columnCount = binarifyColumnSameSheet(wbIn, wbOut, binaryCol, columnCount)
    
print ""    
print "Real-valued Features:"    
print ""
for realCol in LIST_OF_REAL_COLS:
    columnCount = realifyColumnSameSheet(wbIn, wbOut, realCol, columnCount)
    
print ""    
print "Result:"    
copyResultColumn(wbIn, wbOut, ANSWER_COL, columnCount)

wbOut.save(SHEET_NAME[:-5] + "_processed.xlsx")




#binarifyColumn(wb, sys.argv[1])
#wb.save(SHEET_NAME[:-5] + "_binarified_" + sys.argv[1] + ".xlsx")
    
#newWB = openpyxl.Workbook()
#ws = newWB.active
#ws["A1"].value = "hello world"
#newWB.save("hello.xlsx")

    
    

#print vectorizer

#print wb["all"]["A1"].value