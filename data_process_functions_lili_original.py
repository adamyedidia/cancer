import openpyxl, pprint
import numpy as np
import string
import sys
import numbers
import os
import xlrd
import csv
import matplotlib.pyplot as plt
import numpy as np

import xlrd
import csv

from sklearn.feature_extraction.text import CountVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn import cross_validation
from sklearn.metrics import f1_score
from sklearn.metrics import precision_score
from sklearn.metrics import recall_score

from sklearn.pipeline import make_pipeline
from sklearn.metrics import roc_auc_score
from sklearn.metrics import roc_auc_score
from sklearn.ensemble import ExtraTreesClassifier,RandomTreesEmbedding
from sklearn.preprocessing import OneHotEncoder
from sklearn.cross_validation import (cross_val_predict,StratifiedKFold)
from sklearn.metrics import classification_report
import sklearn.cross_validation
import random

from sklearn.metrics import precision_recall_curve
from sklearn.metrics import average_precision_score

from sklearn.pipeline import make_pipeline
from sklearn.metrics import roc_auc_score
from sklearn.metrics import roc_curve,f1_score,precision_score, recall_score
from sklearn.ensemble import ExtraTreesClassifier,RandomTreesEmbedding
from sklearn.preprocessing import OneHotEncoder
from sklearn.cross_validation import (cross_val_predict,StratifiedKFold)
from sklearn.metrics import classification_report
#MAX_ROW = 1556

def convertToColumnLabel(x):

    lastDigit = string.ascii_uppercase[x % 26]

    dividend = x - (x%26)

    if dividend == 0:
        return lastDigit

    else:
        return convertToColumnLabel(dividend/26 - 1) + lastDigit

def addToValueList(differentValueDict, v, i):
    if v != None and v != "None" and v != "":
        if v in differentValueDict:
            differentValueDict[v].append(i)
        else:
            differentValueDict[v] = [i]

def getDifferentValueDict(wb, columnName):
    differentValueDict = {}

    mainSheet = wb.active

    columnDescriptor = mainSheet[columnName + "1"].value.strip()

    print columnDescriptor
    MAX_ROW=mainSheet.max_row

    for i in range(2, mainSheet.max_row+1):
        cellName = columnName + str(i)
        cellValue = str(mainSheet[cellName].value).strip()

        differentValues = string.split(cellValue, ",")

        for v in differentValues:
            addToValueList(differentValueDict, v, i)

    return differentValueDict, columnDescriptor
'''
def binarifyColumn(wb, columnName,MAX_ROW):
    differentValueDict = getDifferentValueDict(wb, columnName)

    newSheet = wb.create_sheet()
    newSheet.title = columnDescriptor
    #MAX_ROW=newSheet.max_row
    for i in range(2, MAX_ROW+1):
        for j in range(len(differentValueDict)):
            newSheet[convertToColumnLabel(j) + str(i)].value = 0

    for j, v in enumerate(differentValueDict.keys()):
        newSheet[convertToColumnLabel(j) + "1"].value = columnDescriptor + "_" + v

        for i in differentValueDict[v]:
            newSheet[convertToColumnLabel(j) + str(i)].value = 1
'''
def binarifyColumnSameSheet(wbIn, wbOut, columnName, columnCount):
    differentValueDict, columnDescriptor = getDifferentValueDict(wbIn, columnName)

    imaginarySheet = {}
    newSheet = wbOut.active
    MAX_ROW=wbIn.active.max_row
    for i in range(2, MAX_ROW+1):
        for j in range(columnCount, columnCount+len(differentValueDict)):
            newSheet[convertToColumnLabel(j) + str(i)].value = 0

    for j, v in enumerate(differentValueDict.keys()):
        newSheet[convertToColumnLabel(j+columnCount) + "1"].value = columnDescriptor + "_" + v

        for i in differentValueDict[v]:
            newSheet[convertToColumnLabel(j+columnCount) + str(i)].value = 1

    return columnCount+len(differentValueDict)
'''
def realifyColumnSameSheet(wbIn, wbOut, columnName, columnCount):
    valueSum = 0
    numValues = 0.

    inSheet = wbIn.active
    outSheet = wbOut.active
    MAX_ROW=wbIn.active.max_row
    columnDescriptor = inSheet[columnName + "1"].value
    print columnDescriptor
    for i in range(1, MAX_ROW+1):
        cellValue = inSheet[columnName + str(i)].value
        try:
            cellValue=float(cellValue.strip())
            #print("yada")
        except:
            pass
        outSheet[convertToColumnLabel(columnCount) + str(i)].value = cellValue
    return columnCount + 1
'''

def realifyColumnSameSheet(wbIn, wbOut, columnName, columnCount):
    MAX_ROW=wbIn.active.max_row
    valueSum = 0
    numValues = 0.

    inSheet = wbIn.active
    outSheet = wbOut.active

    columnDescriptor = inSheet[columnName + "1"].value.strip()
    print columnDescriptor

    for i in range(2, MAX_ROW+1):
        cellValue = inSheet[columnName + str(i)].value
        if isinstance(cellValue, numbers.Number):
            valueSum += cellValue
            numValues += 1.

    averageValue = valueSum/numValues

    outSheet[convertToColumnLabel(columnCount) + "1"].value = columnDescriptor

    for i in range(2, MAX_ROW+1):
        cellValue = inSheet[columnName + str(i)].value
        if isinstance(cellValue, numbers.Number):
            outSheet[convertToColumnLabel(columnCount) + str(i)].value = inSheet[columnName + str(i)].value
        else:
            if cellValue:
                print "Suspicious value", cellValue, "found on row", i, "of column", columnDescriptor + "."
                print "Value ignored."
            outSheet[convertToColumnLabel(columnCount) + str(i)].value = averageValue

    return columnCount + 1

def copyResultColumn(wbIn, wbOut, columnName, columnCount):
    MAX_ROW=wbIn.active.max_row
    inSheet = wbIn.active
    outSheet = wbOut.active
    outSheet[convertToColumnLabel(columnCount) + str(1)].value =inSheet[columnName + "1"].value
    for i in range(2, MAX_ROW+1):
        cellValue = inSheet[columnName + str(i)].value

        outSheet[convertToColumnLabel(columnCount) + str(i)].value=cellValue.strip()
        #print cellValue
        '''
        if cellValue.strip()=='M':
            #print 'yes'
            outSheet[convertToColumnLabel(columnCount) + str(i)].value = 1
        else:
            outSheet[convertToColumnLabel(columnCount) + str(i)].value = 0
        '''
    print inSheet[columnName + "1"].value

    return columnCount + 1


def xls_to_csv(filename):
    os.chdir("C:\Users\liliyu\Dropbox (MIT)\RBG\====000Summer_Project 0000000===\MGH data_3_0718")
    x =  xlrd.open_workbook('filename')
    x1 = x.sheet_by_name('Sheet')
    csvfile = open('data.csv', 'wb')
    writecsv = csv.writer(csvfile, quoting=csv.QUOTE_ALL)
    for rownum in xrange(x1.nrows):
        writecsv.writerow(x1.row_values(rownum))
    csvfile.close()

def read_fromxlsx(SHEET_NAME):
    wbIn = openpyxl.load_workbook(SHEET_NAME)
    sheet = wbIn.active
    maxrow=sheet.max_row
    maxcolumn=sheet.max_column
    header=[]
    for j in range(1,maxcolumn+1):
        header.append(sheet.cell(row=1,column=j).value)
    feature_names=header
    data=[]
    for i in range(2, maxrow+1):
        feats=[]
        for j in range(1,maxcolumn+1):
            feats.append(float(sheet.cell(row=i,column=j).value))
        data.append(feats)
    data=np.array(data)
    return data

def dataprocess(SHEET_NAME,LIST_OF_BINARY_COLS,LIST_OF_REAL_COLS,ANSWER_COL):
    wbIn = openpyxl.load_workbook(SHEET_NAME)
    #MAX_ROW = wbIn.active.max_row
    wbOut = openpyxl.Workbook()
    columnCount = 0

    print "Real-valued Features:"
    for realCol in LIST_OF_REAL_COLS:
        columnCount = realifyColumnSameSheet(wbIn, wbOut, realCol, columnCount)

    print ""
    print "Binary Features:"
    for binaryCol in LIST_OF_BINARY_COLS:
        columnCount = binarifyColumnSameSheet(wbIn, wbOut, binaryCol, columnCount)

    print ""
    print "Result:"
    copyResultColumn(wbIn, wbOut, ANSWER_COL, columnCount)
    wbOut.save(SHEET_NAME[:-5] + "_processed.xlsx")

    #  wbIn = openpyxl.load_workbook(SHEET_NAME)
    sheet = wbOut.active
    maxrow=sheet.max_row
    maxcolumn=sheet.max_column
    print 'data size is: %d X %d' %(maxrow, maxcolumn)
    header=[]
    result=[]
    for j in range(1,maxcolumn):
        header.append(sheet.cell(row=1,column=j).value)
    feature_names=header
    data=[]
    print 'test reasult is: '+str(sheet.cell(row=1,column=maxcolumn).value)
    for i in range(2, maxrow+1):
        feats=[]
        for j in range(1,maxcolumn):
            feats.append(float(sheet.cell(row=i,column=j).value))
        data.append(feats)
        result.append(sheet.cell(row=i,column=maxcolumn).value)
    data=np.array(data)
    return data,feature_names,result


def cus_probability(threshold, model_probability):
    probs=model_probability
    probability=[]
    for p in probs:
        if p>threshold:
            probability.append(1)
        else:
            probability.append(0)
    return probability

def cus_predict (threshold, model_probability):
    probs=model_probability
    predict=[]
    for p in probs:
        if p>threshold:
            predict.append(1)
        else:
            predict.append(0)
    return predict


from sklearn.tree import DecisionTreeClassifier
from sklearn.tree import export_graphviz
from sklearn.grid_search import GridSearchCV
from sklearn.grid_search import RandomizedSearchCV
from sklearn.cross_validation import  cross_val_score


import os
import subprocess

from time import time
from operator import itemgetter
from scipy.stats import randint

import pandas as pd
import numpy as np
from sklearn.tree import DecisionTreeClassifier
from sklearn.tree import export_graphviz
from sklearn.grid_search import GridSearchCV
from sklearn.grid_search import RandomizedSearchCV
from sklearn.cross_validation import  cross_val_score
def report(grid_scores, n_top=3):
    """Report top n_top parameters settings, default n_top=3.
    Args
    ----
    grid_scores -- output from grid or random search
    n_top -- how many to report, of top models
    Returns
    -------
    top_params -- [dict] top parameter settings found in search
    """
    top_scores = sorted(grid_scores,
                        key=itemgetter(1),
                        reverse=True)[:n_top]
    for i, score in enumerate(top_scores):
        print("Model with rank: {0}".format(i + 1))
        print(("Mean validation score: "
               "{0:.3f} (std: {1:.3f})").format(
               score.mean_validation_score,
               np.std(score.cv_validation_scores)))
        print("Parameters: {0}".format(score.parameters))
        print("")

    return top_scores[0].parameters

def run_gridsearch(X, y, clf, param_grid,score, cv=50):
    """Run a grid search for best Decision Tree parameters.

    Args
    ----
    X -- features
    y -- targets (classes)
    cf -- scikit-learn Decision Tree
    param_grid -- [dict] parameter settings to test
    cv -- fold of cross-validation, default 5

    Returns
    -------
    top_params -- [dict] from report()
    """
    grid_search = GridSearchCV(clf,
                               param_grid=param_grid,
                               scoring=score,
                               cv=cv)
    #start = time()
    grid_search.fit(X, y)
    '''
    print(("\nGridSearchCV took {:.2f} "
           "seconds for {:d} candidate "
           "parameter settings.").format(time() - start,
                len(grid_search.grid_scores_)))
    '''
    top_params = report(grid_search.grid_scores_, 10)
    return  top_params


def feature_ranking(forest, feature_names,X,y,outputfile,pp=0):
    forest.fit(X,y)
    #feature_names = vectorizer.get_feature_names()
    importances = forest.feature_importances_
    std = np.std([tree.feature_importances_ for tree in forest.estimators_],
             axis=0)
    indices = np.argsort(importances)[::-1]

    # Print the feature ranking
    print("Feature ranking:")

    output_f=open(outputfile,'wb')
    output_f.write(str(forest))
    output_f.write('\n')
    output_f.write('\n')

    for f in range(X.shape[1]):
        if pp:
            print("%d. feature %s (%f) " % (f + 1,feature_names[indices[f]] , importances[indices[f]]))
        output_f.write("%d. feature %s (%f) " % (f + 1,feature_names[indices[f]] , importances[indices[f]]))
        output_f.write('\n')
    #print("%d. feature %s (%f)  %d  %d" % (f + 1,feature_names[indices[f]] , importances[indices[f]],  M_dic[feature_names[indices[f]]], B_dic[feature_names[indices[f]]]))

    output_f.close()

    # Plot the feature importances of the forest
    plt.figure()
    plt.title("Feature importances")
    #number=X.shape[1]
    number=20
    plt.bar(range(number), importances[indices[0:number]],
       color="r", yerr=std[indices[0:number]], align="center")
    plt.xticks(range(number), indices)
    plt.xlim([-1, number])
    plt.show()


def get_importantfeature(forest,X,y,n):
    forest.fit(X,y)
    #feature_names = vectorizer.get_feature_names()
    importances = forest.feature_importances_
    std = np.std([tree.feature_importances_ for tree in forest.estimators_],
             axis=0)
    indices = np.argsort(importances)[::-1]
    return X[:,indices[:n]]

'''
from sklearn.feature_selection import VarianceThreshold
sel = VarianceThreshold(threshold=(.95 * (1 - .95)))
newf=sel.fit_transform(newft_combined)
newf.shape
'''
def get_sorted_ft_idx(forest,X,y,n):
    forest.fit(X,y)
    #feature_names = vectorizer.get_feature_names()
    importances = forest.feature_importances_
    std = np.std([tree.feature_importances_ for tree in forest.estimators_],
             axis=0)
    indices = np.argsort(importances)[::-1]
    return indices[:n]


### ===========calculate the pearson coeffecient of all features=======
'''
from pydoc import help
from scipy.stats.stats import pearsonr
from collections import defaultdict
pears1=defaultdict(float)
pears2=defaultdict(float)
for i in range(X_structure.shape[1]):
    p=pearsonr(X_structure[:,i],y)
    pears1[feature_names_structure[i]]=p[0]
    pears2[feature_names_structure[i]]=p[1]

import operator
sorted_d = sorted(pears1.items(), key=operator.itemgetter(1))
sorted_dd=[sorted_d[-i-1] for i in range(len(sorted_d))]

f_out=open('pearson on 667 patiens_Reverse.txt','wb')
for pair in sorted_d:
    for a in pair:
        f_out.write(str(a))
        f_out.write(" , ")
    f_out.write('\n')
f_out.close()
'''

# =====================Find the overall faeture ranking=========
'''
from os import listdir
from os.path import isfile, join
import numpy as np
import glob
import glob
from collections import defaultdict
# Reading structure feature
os.chdir('C:\Users\liliyu\Dropbox (MIT)\RBG\====000Summer_Project 0000000===\MGH data_3_0718\report feature')
features_dict=defaultdict(float)
files=glob.glob("*.txt")


for ff in files:
    a=0
    with open(ff,'r') as f:
        while a!='\n':
            a=f.readline()
        a=f.readline()
        while a:
            b=a.split()
            feature=b[2]
            weight=float(b[3].strip('(').strip(')'))
            features_dict[feature]=weight + features_dict[feature]
            #features_dict[feature].append(weight)
            a=f.readline()

import operator
#x = {1: 2, 3: 4, 4: 3, 2: 1, 0: 0}
sorted_d = sorted(features_dict.items(), key=operator.itemgetter(1))
sorted_dd=[sorted_d[-i-1] for i in range(len(sorted_d))]
f_out=open('average_feature_ranking.txt','wb')
for pair in sorted_dd:
    for a in pair:
        f_out.write(str(a))
        f_out.write(" , ")
    f_out.write('\n')
f_out.close()

sorted_dd=[sorted_d[-i-1] for i in range(len(sorted_d))]


f_out=open('average_feature_ranking.txt','wb')
for pair in sorted_dd:
    for a in pair:
        f_out.write(str(a))
        f_out.write(" , ")
    f_out.write('\n')
f_out.close()
'''



'''
import openpyxl
wb = openpyxl.load_workbook('clean_spreadsheet_datawithreports_no assess6 - Copy.xlsx').active
Out = openpyxl.Workbook()
wbOut=Out.active
rowcount=0
for i in range(2, wb.max_row):
    if wb.cell(row=i,column=1).value == wb.cell(row=i-1,column=1).value:
        for j in range(1,wb.max_column):
            wbOut.cell(row=rowcount+1, column=j).value = wb.cell(row=i-1,column=j).value
            wbOut.cell(row=rowcount+2, column=j).value = wb.cell(row=i,column=j).value
        rowcount+=2
Out.save( "duplicates.xlsx")
newdata=[]
newtxt=[]
ID=[]
newy=[]
for i in range(data.shape[0]):
    if data[i,0] not in ID:
        ID.append(data[i,0])
        newdata.append(data[i])
        newtxt.append(ft_report_p[i,:])
        newy.append(yy[i])
    else:
        if yy[i]==1:
            ID.append(data[i,0])
            newdata.append(data[i])
            newtxt.append(ft_report_p[i,:])
            newy.append(yy[i])


newdata=np.array(newdata)
X_structure=newdata[:,1:]
feature_names_structure=feature_names[1:]
y=np.array(newy)
ft_report_prune=np.array(newtxt)
'''

#binarifyColumn(wb, sys.argv[1])
#wb.save(SHEET_NAME[:-5] + "_binarified_" + sys.argv[1] + ".xlsx")

#newWB = openpyxl.Workbook()
#ws = newWB.active
#ws["A1"].value = "hello world"
#newWB.save("hello.xlsx")




#print vectorizer

#print wb["all"]["A1"].value
