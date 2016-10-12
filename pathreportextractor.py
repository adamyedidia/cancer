import openpyxl, pprint
import string

# Contains "Core biopsy" and "breast" but does not contain "excision"

targetSheetName = "reports_column.xlsx"

greatSheetName = "HRL_Full_Data.xlsx"

#numDuplicates = 0

BEFORE_2010 = True

if BEFORE_2010:
    SHEET_NAME = "LehmanPathThru2010.xlsx"
else:
    SHEET_NAME = "LehmanPath2010_2016.xlsx"

wb = openpyxl.load_workbook("../spreadsheets/" + SHEET_NAME)
inSheet = wb.active

greatWB = openpyxl.load_workbook("../spreadsheets/" + greatSheetName)
greatSheet = greatWB.active

targetWB = openpyxl.load_workbook("../spreadsheets/" + greatSheetName)
targetSheet = targetWB.active

if BEFORE_2010:
    REPORT_TEXT_COL = "D"
    MRN_COL = "B"
else:
    REPORT_TEXT_COL = "E"
    MRN_COL = "C"

# s1 contains s2
def containsCaseInsensitive(s1, s2):
    return s2.lower() in s1.lower()

def query(row, column, inSheet):
    entry = inSheet[column + str(row)].value
    
    if type(entry) == type("hello") or type(entry) == type(u"hello"):
        return string.strip(inSheet[column + str(row)].value)
        
    return inSheet[column + str(row)].value
    
def strQuery(row, column, inSheet):
    entry = inSheet[column + str(row)].value
            
    if type(entry) == type("hello") or type(entry) == type(u"hello"):
        return string.strip(inSheet[column + str(row)].value)
        
    return str(inSheet[column + str(row)].value)    

def extractMRN(row, column, inSheet):
    unpaddedNum = strQuery(row, column, inSheet)
    
    return "0"*(7-len(unpaddedNum)) + unpaddedNum

def addToMrnToReportDictWithWarning(mrnToReportDict, mrn, report):
    if mrn in mrnToReportDict:
        print "Warning: duplicate path report for MRN", mrn
        mrnToReportDict[mrn] = ""
        # Should go back and fix these by hand
        
#        numDuplicates += 1
        
    else:
        mrnToReportDict[mrn] = report

if BEFORE_2010:
    MAX_ROW = 641
else:
    MAX_ROW = 863

HRL_MAX_ROW = 1775
HRL_MRN_COL = "A"
TARGET_REPORT_TEXT_COL = "AO"

numReports = 0

mrnToReportDict = {}
mrnToRowDict = {}

# fill out mrnToRowDict
for rowNum in range(1, HRL_MAX_ROW+1):
    mrn = extractMRN(rowNum, HRL_MRN_COL, greatSheet)
    
    mrnToRowDict[mrn] = rowNum

for rowNum in range(1, MAX_ROW+1):
    report = query(rowNum, REPORT_TEXT_COL, inSheet)
    mrn = extractMRN(rowNum, MRN_COL, inSheet)
    
    if containsCaseInsensitive(report, "core biopsy") and \
        containsCaseInsensitive(report, "breast") and \
        (not containsCaseInsensitive(report, "excision")):
        
        addToMrnToReportDictWithWarning(mrnToReportDict, mrn, report)

for mrn in mrnToReportDict:
    if mrn in mrnToRowDict:
        targetRow = mrnToRowDict[mrn]
        report = mrnToReportDict[mrn]
        
        currentTargetSheetValue = targetSheet[TARGET_REPORT_TEXT_COL + str(targetRow)].value
        
        if currentTargetSheetValue != None:
            print "Warning: would overwrite great sheet value for patient in row", targetRow, "ignoring instead"
            
        else:
        
            targetSheet[TARGET_REPORT_TEXT_COL + str(targetRow)].value = report
        
    else:
        print "Warning: no patient found for MRN", mrn
        
targetWB.save("../spreadsheets/" + targetSheetName)        
        
#        numReports += 1

#        print report
#        raw_input()        
        
#print numReports
        
