import openpyxl, pprint
import numpy as np
import string
import sys
import numbers
import re
import random
from sklearn import linear_model
from sklearn.svm import SVC
import matplotlib.pyplot as p

try:
    SHEET_NAME = sys.argv[1]
except:
    raise Exception("\n\nUsage: python word_processing.py [spreadsheet name].xlsx\n" + \
        "Puts the processed spreadsheet in [spreadsheet name]_processed.xlsx")
print "1"
wbIn = openpyxl.load_workbook(SHEET_NAME)
print "2"
inSheet = wbIn.active
print "3"

MAX_ROW = 437

#LIST_OF_BINARY_COLS = ["C", "E", "F", "G", "H", "M", "N", "O", "P", "Q", "R", "S", 
#    "T", "V"]

LIST_OF_BINARY_COLS = []    


    
#LIST_OF_REAL_COLS = ["B", "I", "J", "K", "L"]


ANSWER_COL = "AP"

REPORT_COL = "AO"

FEATURE_EXCLUSION_THRESHOLD = 5

def strSum(s):
    result = ""
    for c in s:
        result += c
    return result

def f1Score(truePositives, trueNegatives, falsePositives, falseNegatives):
#    print truePositives, trueNegatives, falsePositives, falseNegatives

    if truePositives == 0:
        return 0
    
    precision = float(truePositives) / (falsePositives + truePositives)
    recall = float(truePositives) / (truePositives + falseNegatives)
    
    return 2*(precision*recall)/(precision+recall)

def stripOfPunctuation(report):
    newReportList = []
    
    for c in report:
        if not c in string.punctuation:
            newReportList.append(c)
                        
    return "".join(newReportList)
 
def turnDataListIntoArrayAndList(dataList):
    x = np.array([i[0] for i in dataList])
    y = [i[1][0] for i in dataList]
    
    return x,y 

def listOfBigramify(report):
    
    strippedReport = stripOfPunctuation(report)
    words = string.split(strippedReport)
        
    lastWord = words[0]
    
    listOfBigrams = []
    
    for word in words[1:]:  
        listOfBigrams.append((lastWord, word))
        
        lastWord = word
        
    return listOfBigrams

def extractBigrams(sheet):
    bigramDict = {}
    
    for rowNum in range(2, MAX_ROW+1):
        report = sheet[REPORT_COL + str(rowNum)].value
        
        if report:
            listOfBigrams = listOfBigramify(report)
    
            for bigram in listOfBigrams:
                addToNumInstances(bigramDict, bigram)
            
    return bigramDict
    
def pruneRareFeatures(dic, threshold):
    choppingBlock = []
    
    for key in dic:
        if dic[key] <= threshold:
            choppingBlock.append(key)
            
    for key in choppingBlock:
        del dic[key]

def convertToColumnLabel(x):
    lastDigit = string.ascii_uppercase[x % 26]
        
    dividend = x - (x%26)
    
    if dividend == 0:
        return lastDigit
    
    else:
        return convertToColumnLabel(dividend/26 - 1) + lastDigit        

def addToNumInstances(numInstancesDict, v):
    if v in numInstancesDict:
        numInstancesDict[v] += 1
    else:
        numInstancesDict[v] = 1

def addToValueList(differentValueDict, numInstancesDict, v, i):
    if v != None and v != "None" and string.replace(v, " ", "") != "" and v != "OTH" and v != "OTHE" and v != "OTHER":
        if v in differentValueDict:
            differentValueDict[v].append(i)  
            addToNumInstances(numInstancesDict, v)
        else:
            differentValueDict[v] = [i]
            addToNumInstances(numInstancesDict, v)

def getDifferentValueDict(wb, columnName):
    differentValueDict = {}
    numInstancesDict = {}
    
    mainSheet = wb.active
    
    columnDescriptor = mainSheet[columnName + "1"].value
    
    print columnDescriptor
    
    for i in range(2, MAX_ROW+1):
        cellName = columnName + str(i)
        cellValue = str(mainSheet[cellName].value)
        
        differentValues = string.split(cellValue, ",")
        
        for v in differentValues:
            addToValueList(differentValueDict, numInstancesDict, v, i)
            
    deletionList = []        
    for v in differentValueDict:
        if numInstancesDict[v] < FEATURE_EXCLUSION_THRESHOLD:
            print "Deleting feature", columnDescriptor + "_" + v, \
                "with only", str(numInstancesDict[v]), "instances."
            deletionList.append(v)
    
    for item in deletionList:
        del differentValueDict[item]
    
    return differentValueDict, columnDescriptor

def binarifyColumn(wb, columnName):
    differentValueDict = getDifferentValueDict(wb, columnName)
            
    newSheet = wb.create_sheet()
    newSheet.title = columnDescriptor    
    
    for i in range(2, MAX_ROW+1):        
        for j in range(len(differentValueDict)):
            newSheet[convertToColumnLabel(j) + str(i)].value = 0
            
    for j, v in enumerate(differentValueDict.keys()):        
        newSheet[convertToColumnLabel(j) + "1"].value = columnDescriptor + "_" + v
        
        for i in differentValueDict[v]:
            newSheet[convertToColumnLabel(j) + str(i)].value = 1    

def binarifyColumnSameSheet(wbIn, wbOut, columnName, columnCount):
    differentValueDict, columnDescriptor = getDifferentValueDict(wbIn, columnName)
    
    imaginarySheet = {}
    newSheet = wbOut.active
        
    for i in range(2, MAX_ROW+1):        
        for j in range(columnCount, columnCount+len(differentValueDict)):
            newSheet[convertToColumnLabel(j) + str(i)].value = 0
            
    for j, v in enumerate(differentValueDict.keys()):        
        newSheet[convertToColumnLabel(j+columnCount) + "1"].value = columnDescriptor + "_" + v
        
        for i in differentValueDict[v]:
            newSheet[convertToColumnLabel(j+columnCount) + str(i)].value = 1    
            
    return columnCount+len(differentValueDict)
    
def realifyColumnSameSheet(wbIn, wbOut, columnName, columnCount):
    valueSum = 0
    numValues = 0.
        
    inSheet = wbIn.active
    outSheet = wbOut.active
    
    columnDescriptor = inSheet[columnName + "1"].value    
        
    print columnDescriptor    
        
    for i in range(2, MAX_ROW+1):
        cellValue = inSheet[columnName + str(i)].value
        if isinstance(cellValue, numbers.Number):
            valueSum += cellValue
            numValues += 1.
            
    averageValue = valueSum/numValues
    
    outSheet[convertToColumnLabel(columnCount) + "1"].value = columnDescriptor
    
    for i in range(2, MAX_ROW+1):
        cellValue = inSheet[columnName + str(i)].value
        if isinstance(cellValue, numbers.Number):
            outSheet[convertToColumnLabel(columnCount) + str(i)].value = inSheet[columnName + str(i)].value
        else:
            if cellValue:
                print "Suspicious value", cellValue, "found on row", i, "of column", columnDescriptor + "."
                print "Value ignored."
            outSheet[convertToColumnLabel(columnCount) + str(i)].value = averageValue
            
    return columnCount + 1
        
def copyResultColumn(wbIn, wbOut, columnName, columnCount):
    inSheet = wbIn.active
    outSheet = wbOut.active
        
    for i in range(1, MAX_ROW+1):
        cellValue = inSheet[columnName + str(i)].value
        outSheet[convertToColumnLabel(columnCount) + str(i)].value = inSheet[columnName + str(i)].value
    
    print inSheet[columnName + "1"].value
    
    return columnCount + 1

def mainOld():
    wbOut = openpyxl.Workbook()
    columnCount = 0

    print ""
    print "Binary Features:"
    print ""

    for binaryCol in LIST_OF_BINARY_COLS:
        columnCount = binarifyColumnSameSheet(wbIn, wbOut, binaryCol, columnCount)
    
    print ""    
    print "Real-valued Features:"    
    print ""
    for realCol in LIST_OF_REAL_COLS:
        columnCount = realifyColumnSameSheet(wbIn, wbOut, realCol, columnCount)
    
    print ""    
    print "Result:"    
    copyResultColumn(wbIn, wbOut, ANSWER_COL, columnCount)

    wbOut.save(SHEET_NAME[:-5] + "_processed.xlsx")   
   
def vectorify(commonBigrams, inSheet):
    dataList = []
    
    for rowNum in range(2, MAX_ROW+1):
        
        newDataPoint = [[]]
        
        report = inSheet[REPORT_COL + str(rowNum)].value
        
        if report:
            listOfBigrams = listOfBigramify(report)
            
        else:
            listOfBigrams = []    
            
        for bigram in commonBigrams:
            if bigram in listOfBigrams:
                newDataPoint[0].append(1.0)
            else:
                newDataPoint[0].append(0.0)
                
        newDataPoint.append([1.0*("M" in str(inSheet[ANSWER_COL + str(rowNum)].value))])
            
        print 1.0*("M" in str(inSheet[ANSWER_COL + str(rowNum)].value)), inSheet[ANSWER_COL + str(rowNum)].value
            
        dataList.append(newDataPoint)
        
    return dataList

def train(regr, trainingSet):
    x, y = turnDataListIntoTwoArrays(trainingSet)
    regr.fit(x, y)

def printCoeffsForRegression(regr, listOfBigrams):
    listOfCoeffs = []

    for j, coef in enumerate(regr.coef_[0]):
        listOfCoeffs.append((listOfBigrams[j], coef))
        listOfCoeffs.sort(key=lambda x: abs(x[1]))
        listOfCoeffs.reverse()
    
    output = open("coeffs.txt", "w")
    
    for t in listOfCoeffs:
        output.write(str(t[0]) + " " + str(t[1]) + "\n")

def turnDataListIntoTwoArrays(dataList):
    x = np.array([i[0] for i in dataList])
    y = np.array([i[1] for i in dataList])
    
    return x,y
    
def findOptimalThresholds(regr, validationSet):
    x, y = turnDataListIntoArrayAndList(validationSet)

    #predictions = [i[0] for i in regr.predict(x)]
    predictions = regr.predict(x)
    
    truePositiveList = []
    trueNegativeList = []
    falsePositiveList = []
    falseNegativeList = []
    thresholdList = []
    
    for thresholdTimesAHundred in range(-100, 200):
    
        threshold = thresholdTimesAHundred/100.

        ourAnswers = [i > threshold for i in predictions]
    
        truePositives = sum([i and j for i, j in zip(ourAnswers, y)])
        trueNegatives = sum([(not i) and (not j) for i, j in zip(ourAnswers, y)])
        falsePositives = sum([i and (not j) for i, j in zip(ourAnswers, y)])
        falseNegatives = sum([(not i) and j for i, j in zip(ourAnswers, y)])
    
        truePositiveList.append(truePositives)
        trueNegativeList.append(trueNegatives)
        falsePositiveList.append(falsePositives)
        falseNegativeList.append(falseNegatives)
        thresholdList.append(threshold)
    
    f1Scores = [f1Score(tp, tn, fp, fn) for tp, tn, fp, fn in zip(truePositiveList, \
        trueNegativeList, falsePositiveList, falseNegativeList)]    
        
    bestF1threshold = thresholdList[f1Scores.index(max(f1Scores))]

    falseNegativeList.reverse()
    bestNoCancerThresholdReversed = falseNegativeList.index(0)
    falseNegativeList.reverse()
    bestNoCancerThreshold = thresholdList[len(thresholdList) - bestNoCancerThresholdReversed - 1]
    
    print f1Scores
    
    print bestF1threshold, bestNoCancerThreshold
    return bestF1threshold, bestNoCancerThreshold

def test(regr, testSet, threshold):
    x, y = turnDataListIntoArrayAndList(testSet)
#    predictions = [i[0] for i in regr.predict(x)]
    predictions = regr.predict(x)
            
    ourAnswers = [i > threshold for i in predictions]
    
    truePositives = sum([i and j for i, j in zip(ourAnswers, y)])
    trueNegatives = sum([(not i) and (not j) for i, j in zip(ourAnswers, y)])
    falsePositives = sum([i and (not j) for i, j in zip(ourAnswers, y)])
    falseNegatives = sum([(not i) and j for i, j in zip(ourAnswers, y)])

    return truePositives, trueNegatives, falsePositives, falseNegatives    
    
def main():
    bigramDict = extractBigrams(inSheet)
    
    pruneRareFeatures(bigramDict, FEATURE_EXCLUSION_THRESHOLD)
    
    commonBigrams = bigramDict.keys()
    
    dataSet = vectorify(commonBigrams, inSheet)
    
    random.shuffle(dataSet)
    
#    for i in dataSet:
#        print "horkley"
#        if i[1][0] != 0:
#            print i[1]
    
    numPoints = len(dataSet)
    trainingSet = dataSet[:numPoints/2]
    validationSet = dataSet[numPoints/2:3*numPoints/4]
    testSet = dataSet[3*numPoints/4:]

#    regr = linear_model.Ridge(alpha=100.0)
    regr = SVC(kernel="linear")
    
    train(regr, trainingSet)

#    printCoeffsForRegression(regr, commonBigrams)

    train(regr, trainingSet)
    bestF1threshold, bestNoCancerThreshold = findOptimalThresholds(regr, validationSet)

    tp, tn,fp,fn = test(regr, testSet, -0.005)
    
    print "Fraction of classified patients", float(tn + fn) / (tp+fp+fn+tn)
    
    print "Miss rate", float(fn)/float(tp + fn)

    tp, tn,fp,fn = test(regr, testSet, -0.005)
    
    print "Fraction of classified patients", float(tn + fn) / (tp+fp+fn+tn)
    
    print "Miss rate", float(fn)/float(tp + fn)

    tp, tn,fp,fn = test(regr, testSet, -0.01)
    
    print "Fraction of classified patients", float(tn + fn) / (tp+fp+fn+tn)
    
    print "Miss rate", float(fn)/float(tp + fn)

    tp, tn,fp,fn = test(regr, testSet, 0.0)
    
    print "Fraction of classified patients", float(tn + fn) / (tp+fp+fn+tn)
    
    print "Miss rate", float(fn)/float(tp + fn)

    tp, tn,fp,fn = test(regr, testSet, -0.015)
    
    print "Fraction of classified patients", float(tn + fn) / (tp+fp+fn+tn)
    
    print "Miss rate", float(fn)/float(tp + fn)


    tpF1, tnF1, fpF1, fnF1 = test(regr, testSet, bestF1threshold)
    tpNC, tnNC, fpNC, fnNC = test(regr, testSet, bestNoCancerThreshold)
    
    print "Fraction of classified patients", float(tnF1 + fnF1) / (tpF1+fpF1+fnF1+tnF1)
    
    print "Miss rate", float(fnF1)/float(tpF1 + fnF1)

    print "Fraction of classified patients", float(tnNC + fnNC) / (tpNC+fpNC+fnNC+tnNC)
    
    print "Miss rate", float(fnNC)/float(tpNC + fnNC)

    print "Using optimal F1 threshold of", bestF1threshold, "over validation set:"
    print "F1 SCORE:", f1Score(tpF1, tnF1, fpF1, fnF1)
    print "True positives:", tpF1
    print "True negatives:", tnF1
    print "False positives:", fpF1
    print "False negatives:", fnF1
    print ""
    print "Using optimal no-cancer threshold of", bestNoCancerThreshold, "over validation set:"
    print "F1 SCORE:", f1Score(tpNC, tnNC, fpNC, fnNC)
    print "True positives:", tpNC
    print "True negatives:", tnNC
    print "False positives:", fpNC
    print "False negatives:", fnNC    
    
    print len(dataSet)
    
if __name__ == "__main__":    
    main()



#binarifyColumn(wb, sys.argv[1])
#wb.save(SHEET_NAME[:-5] + "_binarified_" + sys.argv[1] + ".xlsx")
    
#newWB = openpyxl.Workbook()
#ws = newWB.active
#ws["A1"].value = "hello world"
#newWB.save("hello.xlsx")

    
    

#print vectorizer

#print wb["all"]["A1"].value