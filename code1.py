# cd "/Users/NINI/Dropbox (MIT)/RBG/====000Summer_Project 0000000===/mgh data"
import openpyxl, pprint
import numpy
from sklearn.feature_extraction.text import CountVectorizer

wb = openpyxl.load_workbook('../spreadsheets/data_pruned_h_only.xlsx')
sheet = wb.get_sheet_by_name('all')
vectorizer = CountVectorizer(min_df=1)
sheet = wb.active

feature=numpy.empty([4854,1])
#for i in range(sheet.max_column):
for i in range(sheet.max_column):
    feat=sheet.columns[i]
    hist=[]
    for i in range(len(feat)):
        if feat[i].value:
            cc=str(feat[i].value).split(',')
            c=' '.join(cc)
            #print c
        else: 
            c='NNN'
        hist.append(c)
    X = vectorizer.fit_transform(hist)
    xx=X.toarray()
    xx=numpy.array(xx)
    feature=numpy.column_stack((feature,xx))

target=[]
result=sheet.columns[sheet.max_column-1]
for i in range(len(result)):
    target.append(result[i].value)
    
from sklearn.linear_model import LogisticRegression

xtrain=feature[1:4000]
ytrain=target[1:4000]
xtest=feature[4000:]
ytest=target[4000:]

clf = LogisticRegression()
clf.fit(xtrain,ytrain)

clf.score(xtest,ytest)